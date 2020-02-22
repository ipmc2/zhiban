from openpyxl import load_workbook
import datetime
import pyperclip
from time import sleep

wb = load_workbook(r'./值班人员/值班人员.xlsx')
sheet = wb.worksheets[0]
rows = sheet.max_row
cloums = sheet.max_column
count = 3
col_value_str = ''
for j in range(3, rows):
    col_value_str = sheet.cell(row=j, column=1).value
    timezone = col_value_str.timestamp() - datetime.datetime.now().timestamp()
    timezone_conver = int(timezone/3600)
    if (timezone_conver > -7) and (timezone_conver < 7):
        break
    count += 1
date = str(col_value_str).split()
if int(date[1].split(':')[0]) > 10:
    clas = '夜班'
else:
    clas = '日班'
title = str(date[0]) + '十四及知识城线' + clas+'值班人员安排:' + '\n'
info1 = ''
for i in range(1, cloums):
   info1 = info1 + str(sheet.cell(row=1, column=i+1).value) + ':' + str(sheet.cell(row=count, column=i+1).value) + '\n'
allinfo = title + info1
pyperclip.copy(allinfo)
print(allinfo)
sleep(10)